from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

##idfsup = load_workbook('IDF_Information.xlsx')
##idfactive = idfsup.active
##idfamt = max(idfactive['A'].value)
##print(idfamt)
##future tool info^^^


print('Insert IDF amount')
idfamt = input()
idfamt = int(idfamt)

for eachIDF in range(1, idfamt +1):

	wb = Workbook()

	ws = wb.active
	ws.title = "Information"

	print('Insert IDF Two Digit #')
	idf = input()
	print('Insert # of access switches')
	acc = input()
	print('Insert # of RSWs')
	rsw = input()

	acc = int(acc)
	rsw = int(rsw)

	ws.append(['Start','Location','End', 'Location', 'Type'])
	ws.append(['bfl2-co-dis-sw'+ idf +'01 Te1/0/39', "IDF-" + idf + ' R-38', "Card 1 Pair 1", "IDF-" + idf + " R-41", "Single Mode Fiber Yellow 1m"])
	ws.append(['bfl2-co-dis-sw'+ idf +'01 Te1/0/40', "IDF-" + idf + ' R-38', "Card 2 Pair 1", "IDF-" + idf + " R-41", "Single Mode Fiber Yellow 1m"])
	ws.append(['bfl2-co-dis-sw'+ idf +'02 Te1/0/39', "IDF-" + idf + ' R-37', "Card 1 Pair 2", "IDF-" + idf + " R-41", "Single Mode Fiber Yellow 1m"])
	ws.append(['bfl2-co-dis-sw'+ idf +'02 Te1/0/40', "IDF-" + idf + ' R-37', "Card 2 Pair 2", "IDF-" + idf + " R-41", "Single Mode Fiber Yellow 1m"])
	ws.append(['bfl2-co-con-sw'+ idf +'01 Gi1/0/1', 'IDF-' + idf + ' R-39', 'Card 1 Pair 3', 'IDF-' + idf + ' R-41', 'Single Mode Fiber Green 3m'])
	ws.append(['bfl2-co-con-sw'+ idf +'01 Port 1', 'IDF-' + idf + ' R-39', 'bfl2-co-dis-sw'+ idf +'01 Con. Port', 'IDF-' + idf + ' R-38', 'Green Rollover 3ft'])
	ws.append(['bfl2-co-con-sw'+ idf +'01 Port 2', 'IDF-' + idf + ' R-39', 'bfl2-co-dis-sw'+ idf +'02 Con. Port', 'IDF-' + idf + ' R-37', 'Green Rollover 3ft'])


	for x in range(1, acc + 1):
		ws.append(['bfl2-co-dis-sw'+ idf +'01 Te1/0/' + str(x), "IDF-" + idf + ' R-38', "bfl2-co-acc-sw" + idf + '0' + str(x) + ' Te1/1/7', "IDF-" + idf + ' R-' + str(36-3*x), "Single Mode Fiber Yellow 1m"])

	for f in range (1, rsw + 1):
		ws.append(['bfl2-co-dis-sw'+ idf +'01 Te1/0/' + str(f+31), "IDF-" + idf + ' R-38', "bfl2-co-acc-rsw" + str(int(float(idf)))+'-'+str(f)+' Te1/1/7', "IDF-" + idf + ' R-' + str(4+3*f), "Single Mode Fiber Yellow 1m"])


	for x in range (1, acc + 1):
		ws.append(['bfl2-co-dis-sw'+ idf +'02 Te1/0/' + str(x), "IDF-" + idf + ' R-37', "bfl2-co-acc-sw" + idf + '0' + str(x) + ' Te1/1/8', "IDF-" + idf + ' R-' + str(36-3*x), "Single Mode Fiber Yellow 1m"])

	for f in range (1, rsw + 1):
		ws.append(['bfl2-co-dis-sw'+ idf +'02 Te1/0/' + str(f+31), "IDF-" + idf + ' R-37', "bfl2-co-acc-rsw" + str(int(float(idf)))+'-'+str(f)+' Te1/1/8', "IDF-" + idf + ' R-' + str(4+3*f), "Single Mode Fiber Yellow 1m"])

	for z in range (1, acc + 1):
		ws.append(['bfl2-co-con-sw'+ idf + '01 Port ' + str(z+2), 'IDF-' + idf + ' R-39', 'bfl2-co-acc-sw' + idf + '0' + str(z) + ' Con. Port', 'IDF-' + idf + ' R-' + str(36-3*z), 'Green Rollover 3ft'])

	for y in range (1, rsw + 1):
		ws.append(['bfl2-co-con-sw'+ idf + '01 Port ' + str(y+11), 'IDF-' + idf + ' R-39', 'bfl2-co-acc-rsw' + str(int(float(idf)))+'-'+str(y)+' Con. Port', 'IDF-' + idf + ' R-' + str(4+3*y), 'Green Rollover 3ft'])


	wb.save(idf + 'Labels.xlsx')
